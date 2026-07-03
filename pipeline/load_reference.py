"""Load severity weights, the LSOA->Ward->LAD lookup, and ONS populations.

The two ONS xlsx files are the only inputs whose internal layout we could not
verify up front, so both parsers locate their header rows dynamically and fail
loudly (with what they saw) if the shape is not what we expect.
"""
from __future__ import annotations

import json
import re

import duckdb
from openpyxl import load_workbook

import config

LSOA_CODE_RE = re.compile(r"^[EW]01\d{6}$")
LAD_CODE_RE = re.compile(r"^(E0[6789]|W06|N09|S12)\d{6}$")


def load_severity(con: duckdb.DuckDBPyConnection) -> None:
    con.execute("CREATE OR REPLACE TABLE severity (category VARCHAR PRIMARY KEY, weight DOUBLE, tier INTEGER)")
    con.executemany(
        "INSERT INTO severity VALUES (?, ?, ?)",
        [(cat, w, t) for cat, (w, t) in config.SEVERITY.items()],
    )


def load_lookup(con: duckdb.DuckDBPyConnection) -> None:
    rows = json.loads((config.RAW_DIR / "lookup_lsoa.json").read_text())
    con.execute(
        "CREATE OR REPLACE TABLE lookup_lsoa ("
        " lsoa21cd VARCHAR PRIMARY KEY, lsoa21nm VARCHAR,"
        " wd25cd VARCHAR, wd25nm VARCHAR, lad25cd VARCHAR, lad25nm VARCHAR)"
    )
    con.executemany(
        "INSERT INTO lookup_lsoa VALUES (?, ?, ?, ?, ?, ?)",
        [
            (r["LSOA21CD"], r["LSOA21NM"], r["WD25CD"], r["WD25NM"], r["LAD25CD"], r["LAD25NM"])
            for r in rows
        ],
    )
    n = con.execute("SELECT count(*) FROM lookup_lsoa").fetchone()[0]
    print(f"lookup_lsoa: {n} rows")


def _find_header(rows: list[tuple], code_pattern: str, value_pattern: str) -> tuple[int, int, int]:
    """Scan the first rows for a header row; return (row_idx, code_col, value_col)."""
    code_re = re.compile(code_pattern, re.I)
    value_re = re.compile(value_pattern, re.I)
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        code_col = next((j for j, c in enumerate(cells) if code_re.search(c)), None)
        value_col = next((j for j, c in enumerate(cells) if value_re.search(c)), None)
        if code_col is not None and value_col is not None:
            return i, code_col, value_col
    raise RuntimeError(
        f"header row not found (code~/{code_pattern}/, value~/{value_pattern}/); "
        f"first rows: {[r[:8] for r in rows[:6]]}"
    )


def _pick_sheet(sheetnames: list[str], must: list[str], prefer: list[str]) -> str:
    candidates = [s for s in sheetnames if all(re.search(m, s, re.I) for m in must)]
    if not candidates:
        raise RuntimeError(f"no sheet matching {must} in {sheetnames}")
    for p in prefer:
        preferred = [s for s in candidates if re.search(p, s, re.I)]
        if preferred:
            candidates = preferred
    return candidates[0]


def parse_lsoa_population() -> list[tuple[str, int]]:
    wb = load_workbook(config.RAW_DIR / "pop_lsoa.xlsx", read_only=True, data_only=True)
    try:
        sheet = _pick_sheet(wb.sheetnames, must=[r"mid[ -]?2024"], prefer=[r"lsoa"])
        ws = wb[sheet]
        head = [r for _, r in zip(range(12), ws.iter_rows(values_only=True))]
        row_idx, code_col, total_col = _find_header(head, r"LSOA.*code", r"^total")
        out = []
        for row in ws.iter_rows(min_row=row_idx + 2, values_only=True):
            code = str(row[code_col]).strip() if row[code_col] else ""
            if LSOA_CODE_RE.match(code) and isinstance(row[total_col], (int, float)):
                out.append((code, int(row[total_col])))
        total = sum(p for _, p in out)
        if not (30_000 < len(out) < 40_000 and 55e6 < total < 65e6):
            raise RuntimeError(f"LSOA population implausible: {len(out)} rows, total {total:,}")
        print(f"population/lsoa: sheet '{sheet}', {len(out)} rows, total {total:,}")
        return out
    finally:
        wb.close()


def parse_lad_population() -> list[tuple[str, int]]:
    wb = load_workbook(config.RAW_DIR / "pop_lad.xlsx", read_only=True, data_only=True)
    try:
        sheet = _pick_sheet(wb.sheetnames, must=[r"MYE\s*2"], prefer=[r"person"])
        ws = wb[sheet]
        head = [r for _, r in zip(range(12), ws.iter_rows(values_only=True))]
        row_idx, code_col, total_col = _find_header(head, r"^code$", r"all\s*ages")
        out = []
        for row in ws.iter_rows(min_row=row_idx + 2, values_only=True):
            code = str(row[code_col]).strip() if row[code_col] else ""
            if LAD_CODE_RE.match(code) and isinstance(row[total_col], (int, float)):
                out.append((code, int(row[total_col])))
        by_prefix = {p: sum(1 for c, _ in out if c.startswith(p)) for p in ("E0", "W06", "N09", "S12")}
        if by_prefix["N09"] != 11 or len(out) < 350:
            raise RuntimeError(f"LAD population implausible: {len(out)} rows, by prefix {by_prefix}")
        pop = dict(out)
        for old, new in config.LAD_RECODES.items():
            if old in pop and new not in pop:
                pop[new] = pop.pop(old)
        print(f"population/lad: sheet '{sheet}', {len(pop)} rows, by prefix {by_prefix}")
        return list(pop.items())
    finally:
        wb.close()


def load_population(con: duckdb.DuckDBPyConnection) -> None:
    lsoa = parse_lsoa_population()
    lad = parse_lad_population()
    con.execute("CREATE OR REPLACE TABLE population (code VARCHAR, level VARCHAR, population BIGINT)")
    con.executemany("INSERT INTO population VALUES (?, 'lsoa', ?)", lsoa)
    con.executemany("INSERT INTO population VALUES (?, 'lad', ?)", lad)
    # Ward population (E+W) = best-fit sum of member LSOA populations.
    con.execute(
        """
        INSERT INTO population
        SELECT l.wd25cd, 'ward', sum(p.population)
        FROM lookup_lsoa l JOIN population p ON p.code = l.lsoa21cd AND p.level = 'lsoa'
        GROUP BY l.wd25cd
        """
    )
    counts = dict(con.execute("SELECT level, count(*) FROM population GROUP BY level").fetchall())
    print(f"population table: {counts}")


def run(con: duckdb.DuckDBPyConnection) -> None:
    load_severity(con)
    load_lookup(con)
    load_population(con)


if __name__ == "__main__":
    with duckdb.connect(str(config.DB_PATH)) as con:
        run(con)
